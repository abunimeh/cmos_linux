# Author: Guanyu Yi @ CPU Verification Platform Group
# Email: yigy@cpu.com.cn
# Description: VPlanProc class for processing vplan

### modules
import pcom
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import os
import re
from urllib import request
import datetime as dt
import requests
import collections

### classes
class VPlanProc(object):
    def __init__(self, LOG, ced, cfg_dic, days):
        self.LOG = LOG if LOG else pcom.gen_logger()
        self.ced = ced
        self.cfg_dic = cfg_dic
        self.days = days
        self.proj_cfg = self.cfg_dic['proj']
        self.case_cfg = self.cfg_dic['case']
        self.home_row_lst = pcom.rd_cfg(
            self.proj_cfg, 'vplan_sheets', 'home')
        self.home_row_lst.insert(5, '')
        self.tc_col_lst = pcom.rd_cfg(
            self.proj_cfg, 'vplan_sheets', 'test_case')
        self.tc_index = self.tc_col_lst.index('Case Name')
        self.tp_index = self.tc_col_lst.index('Priority')
        self.to_index = self.tc_col_lst.index('Owner')
        self.ts_index = self.tc_col_lst.index('Status')
        self.tr_index = self.tc_col_lst.index('Days')
        self.tv_index = self.tc_col_lst.index('CL Ver')
        self.td_index = self.tc_col_lst.index('Description')
        self.cc_col_lst = pcom.rd_cfg(
            self.proj_cfg, 'vplan_sheets', 'code_coverage')
        self.ch_index = self.cc_col_lst.index('Hierarchy')
        self.cp_index = self.cc_col_lst.index('Priority')
        self.co_index = self.cc_col_lst.index('Owner')
        self.cs_index = self.cc_col_lst.index('Score')
        self.cl_index = self.cc_col_lst.index('Line')
        self.cc_index = self.cc_col_lst.index('Cond')
        self.ct_index = self.cc_col_lst.index('Toggle')
        self.cf_index = self.cc_col_lst.index('FSM')
        self.cb_index = self.cc_col_lst.index('Branch')
        self.ca_index = self.cc_col_lst.index('Assert')
        self.fc_col_lst = pcom.rd_cfg(
            self.proj_cfg, 'vplan_sheets', 'function_coverage')
        self.fc_index = self.fc_col_lst.index('Coverage Group')
        self.fp_index = self.fc_col_lst.index('Priority')
        self.fs_index = self.fc_col_lst.index('SNPS Cov Per')
        self.fd_index = self.fc_col_lst.index('Description')
        self.vplan_file = (self.ced['MODULE_VPLAN']+os.sep+
                           '{0}_vplan.xlsx'.format(self.ced['MODULE']))
        self.gen_flg = False if os.path.isfile(self.vplan_file) else True
        self.tpn = 0
        self.ttn = 0
        self.ccs = '0'
        self.fcp = '0'
    def gen_per_color(self, per):
        if not per.replace('.', '').isdigit():
            return 'FFFFFFFF'
        per = float(per)
        if 0<=per<5:
            cs = 'FFFF0000'
        elif 5<=per<20:
            cs = 'FFFF3C3C'
        elif 20<=per<40:
            cs = 'FFFF2A00'
        elif 40<=per<60:
            cs = 'FFFFA500'
        elif 60<=per<80:
            cs = 'FFAEFF00'
        elif 80<=per<95:
            cs = 'FF04FF00'
        elif 95<=per<=100:
            cs = 'FF00FF00'
        else:
            cs = 'FFFF0000'
        return cs
    def fill_cc_scores(self, cell, score):
        cell.value = '{0} %'.format(score)
        cell.fill = PatternFill(
            fill_type='gray125', end_color=self.gen_per_color(score))
    def parse_ch_report(self):
        cd_rpt_file = self.ced[
            'COV_MERGE_DIR']+os.sep+'urgReport'+os.sep+'dashboard.txt'
        ch_rpt_file = self.ced[
            'COV_MERGE_DIR']+os.sep+'urgReport'+os.sep+'hierarchy.txt'
        if not os.path.isfile(cd_rpt_file):
            raise Exception(
                "merged code coverage dashboard file {0} is NA".format(
                    cd_rpt_file))
        if not os.path.isfile(ch_rpt_file):
            raise Exception(
                "merged code coverage report file {0} is NA".format(
                    ch_rpt_file))
        with open(cd_rpt_file) as rf:
            cd_rpt_con = rf.read()
        self.ccs = re.search(r'Total Coverage Summary.*?(\d+\.\d+)',
                             cd_rpt_con, flags=re.DOTALL).group(1)
        ch_score_dic = collections.OrderedDict()
        with open(ch_rpt_file) as rf:
            ch_rpt_con = rf.read()
        con_lst = re.findall(
            r'{0}\s{{0,2}}-{{10,}}{0}(.*?)(?={0}\s+-{{10}}|$)'.format(
                os.linesep), ch_rpt_con, flags=re.DOTALL)
        for index, con in enumerate(con_lst):
            p_str = '(top)' if index == 0 else ''
            for line in con.split(os.linesep):
                line = line.strip()
                m = pcom.REOpter(line)
                if m.match(re.compile(
                        r'([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+'
                        r'([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+'
                        r'([\d\.\-]+)\s+(\w+)')):
                    ch_score_dic[m.group(8)+p_str] = {
                        's': m.group(1), 'l': m.group(2), 'c': m.group(3),
                        't': m.group(4), 'f': m.group(5), 'b': m.group(6),
                        'a': m.group(7)}
        return ch_score_dic
    def parse_cg_report(self):
        cg_rpt_file = self.ced[
            'COV_MERGE_DIR']+os.sep+'urgReport'+os.sep+'groups.txt'
        cp_rpt_file = self.ced[
            'COV_MERGE_DIR']+os.sep+'urgReport'+os.sep+'grpinfo.txt'
        if not os.path.isfile(cg_rpt_file):
            raise Exception(
                "merged coverage groups report file {0} is NA".format(
                    cg_rpt_file))
        if not os.path.isfile(cp_rpt_file):
            raise Exception(
                "merged coverage points report file {0} is NA".format(
                    cp_rpt_file))
        cg_score_dic = collections.OrderedDict()
        with open(cg_rpt_file) as rf:
            for line in rf:
                line = line.strip()
                m = pcom.REOpter(line)
                if m.match(re.compile(r'(\d+\.\d+)\s+\d+$')):
                    self.fcp = m.group(1)
                elif m.match(re.compile(r'(\d+\.\d+)\s+.*\w+::\w+::(\w+)')):
                    cg_score_dic[m.group(2)] = {
                        'per': m.group(1), 'cp_dic': collections.OrderedDict()}
        with open(cp_rpt_file) as rf:
            cp_rpt_con = rf.read()
        for cg, cg_dic in cg_score_dic.items():
            cg_sum_con = re.search(
                r'Summary for Group\s+(?:\w+::)+{0}(.*?{1}-{{60}})'.format(
                    cg, os.linesep), cp_rpt_con, flags=re.DOTALL).group(1)
            var_con, cro_con = re.search(
                r'Variables for Group\s+(?:\w+::)+{0}(.*?){1}Crosses for Group'
                r'\s+(?:\w+::)+{0}(.*?){1}-{{60}}'.format(cg, os.linesep),
                cg_sum_con, flags=re.DOTALL).groups()
            for line in var_con.split(os.linesep):
                line = line.strip()
                m = pcom.REOpter(line)
                if m.match(re.compile(
                        r'(\w+)\s+(?:\d+\s+)+(\d+\.\d+)\s+(?:\d+\s+)+')):
                    cg_dic['cp_dic'][cg+'::'+m.group(1)] = m.group(2)
            for line in cro_con.split(os.linesep):
                line = line.strip()
                m = pcom.REOpter(line)
                if m.match(re.compile(
                        r'(\w+)\s+(?:\d+\s+)+(\d+\.\d+)\s+(?:\d+\s+)+')):
                    cg_dic['cp_dic'][cg+'::'+m.group(1)+'(cross)'] = m.group(2)
        return cg_score_dic
    def proc_home_sheet(self, ws):
        ws.title = 'home'
        for index, row in enumerate(self.home_row_lst):
            if index == 5:
                continue
            cell = ws['a'+str(index+1)]
            cell.value = row
            if self.gen_flg:
                cell.style = 'Accent1'
            cell.alignment = Alignment(wrap_text=True)
            next_cell = ws['b'+str(index+1)]
            if row == 'Project':
                next_cell.value = self.ced['PROJ_NAME']
            elif row == 'Module Name':
                next_cell.value = self.ced['MODULE']
            elif row == 'Case Passing Rate':
                dv = self.tpn/self.ttn if self.ttn else 0
                cpr = str(round(100*dv, 2))
                next_cell.value = '{0} % ({1}/{2})'.format(
                    cpr, self.tpn, self.ttn)
                next_cell.fill = PatternFill(
                    fill_type='gray125', end_color=self.gen_per_color(cpr))
            elif row == 'Code Coverage Score':
                next_cell.value = '{0} %'.format(self.ccs)
                next_cell.fill = PatternFill(
                    fill_type='gray125', end_color=self.gen_per_color(
                        self.ccs))
            elif row == 'Function Coverage Per':
                next_cell.value = '{0} %'.format(self.fcp)
                next_cell.fill = PatternFill(
                    fill_type='gray125', end_color=self.gen_per_color(
                        self.fcp))
        ws.column_dimensions['a'].width = pcom.rd_cfg(
            self.proj_cfg, 'vplan_column_width', ws.title)[0]
        ws.column_dimensions['b'].width = ws.column_dimensions['a'].width
    def proc_tc_sheet(self, ws):
        query_url = 'http://172.51.13.205:8000/regr/db_query/query_case_dic/'
        query_param = {'date': dt.datetime.now().strftime('%Y_%m_%d'),
                       'proj': self.ced['PROJ_NAME'],
                       'module': self.ced['MODULE'],
                       'days': self.days}
        r = requests.get(query_url, params=query_param)
        case_pr_dic = r.json()
        if self.gen_flg:
            ws.append(self.tc_col_lst)
        for cell, width in zip(ws[1], pcom.rd_cfg(
                self.proj_cfg, 'vplan_column_width', ws.title)):
            if self.gen_flg:
                cell.style = 'Accent1'
            cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions[cell.column].width = width
        for index, case_row in enumerate(ws.rows):
            if index == 0:
                continue
            case_name = case_row[self.tc_index].value
            case_dic = case_pr_dic.get(case_name, {})
            if case_name in self.case_cfg:
                case_row[self.td_index].value = self.case_cfg[
                    case_name]['vplan_desc'].replace(os.linesep, '; ')
                case_row[self.tp_index].value = self.case_cfg[
                    case_name]['vplan_priority']
                case_row[self.to_index].value = self.case_cfg[
                    case_name]['vplan_owner']
                case_row[self.ts_index].value = '{0} % ({1}/{2})'.format(
                    str(case_dic.get('pr', 0.0)), str(case_dic.get('pn', 0)),
                    str(case_dic.get('tn', 0)))
                self.tpn += case_dic.get('pn', 0)
                self.ttn += case_dic.get('tn', 0)
                case_row[self.ts_index].fill = PatternFill(
                    fill_type='gray125',
                    end_color=case_dic.get('bc', '#FF0000').replace('#', 'FF'))
                case_row[self.tr_index].value = self.days
                case_row[self.tv_index].value = case_dic.get('cl_range', 'NA')
                self.case_cfg.pop(case_name)
            else:
                case_row[self.tp_index].value = 'Out of Date'
                case_row[self.tp_index].fill = PatternFill(
                    fill_type='gray125', end_color='FFFF0000')
        for case_name in self.case_cfg:
            if case_name == 'DEFAULT':
                continue
            case_dic = case_pr_dic.get(case_name, {})
            new_line = ['']*len(self.tc_col_lst)
            new_line[self.tc_index] = case_name
            new_line[self.td_index] = self.case_cfg[
                case_name]['vplan_desc'].replace(os.linesep, '; ')
            new_line[self.tp_index] = self.case_cfg[
                case_name]['vplan_priority']
            new_line[self.to_index] = self.case_cfg[
                case_name]['vplan_owner']
            new_line[self.ts_index] = '{0} % ({1}/{2})'.format(
                str(case_dic.get('pr', 0.0)),
                str(case_dic.get('pn', 0)),
                str(case_dic.get('tn', 0)))
            self.tpn += case_dic.get('pn', 0)
            self.ttn += case_dic.get('tn', 0)
            new_line[self.tr_index] = self.days
            new_line[self.tv_index] = case_dic.get('cl_range', 'NA')
            ws.append(new_line)
            ws[ws._current_row][self.ts_index].fill = PatternFill(
                fill_type='gray125',
                end_color=case_dic.get('bc', '#FF0000').replace('#', 'FF'))
    def proc_cc_sheet(self, ws):
        ch_score_dic = self.parse_ch_report()
        if self.gen_flg:
            ws.append(self.cc_col_lst)
        for cell, width in zip(ws[1], pcom.rd_cfg(
                self.proj_cfg, 'vplan_column_width', ws.title)):
            if self.gen_flg:
                cell.style = 'Accent1'
            cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions[cell.column].width = width
        for index, ch_row in enumerate(ws.rows):
            if index == 0:
                continue
            ch_name = ch_row[self.ch_index].value
            if ch_name in ch_score_dic:
                self.fill_cc_scores(ch_row[self.cs_index],
                                    ch_score_dic[ch_name]['s'])
                self.fill_cc_scores(ch_row[self.cl_index],
                                    ch_score_dic[ch_name]['l'])
                self.fill_cc_scores(ch_row[self.cc_index],
                                    ch_score_dic[ch_name]['c'])
                self.fill_cc_scores(ch_row[self.ct_index],
                                    ch_score_dic[ch_name]['t'])
                self.fill_cc_scores(ch_row[self.cf_index],
                                    ch_score_dic[ch_name]['f'])
                self.fill_cc_scores(ch_row[self.cb_index],
                                    ch_score_dic[ch_name]['b'])
                self.fill_cc_scores(ch_row[self.ca_index],
                                    ch_score_dic[ch_name]['a'])
                ch_score_dic.pop(ch_name)
            else:
                ch_row[self.fp_index].value = 'Out of Date'
                ch_row[self.fp_index].fill = PatternFill(
                    fill_type='gray125', end_color='FFFF0000')
        for ch_name, ch_dic in ch_score_dic.items():
            new_line = ['']*len(self.cc_col_lst)
            new_line[self.ch_index] = ch_name
            new_line[self.cs_index] = '{0} %'.format(ch_dic['s'])
            new_line[self.cl_index] = '{0} %'.format(ch_dic['l'])
            new_line[self.cc_index] = '{0} %'.format(ch_dic['c'])
            new_line[self.ct_index] = '{0} %'.format(ch_dic['t'])
            new_line[self.cf_index] = '{0} %'.format(ch_dic['f'])
            new_line[self.cb_index] = '{0} %'.format(ch_dic['b'])
            new_line[self.ca_index] = '{0} %'.format(ch_dic['a'])
            ws.append(new_line)
            if '(top)' in ch_name:
                ws[ws._current_row][self.ch_index].fill = PatternFill(
                    fill_type='gray125', end_color='FFFFFF00')
            ws[ws._current_row][self.cs_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['s']))
            ws[ws._current_row][self.cl_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['l']))
            ws[ws._current_row][self.cc_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['c']))
            ws[ws._current_row][self.ct_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['t']))
            ws[ws._current_row][self.cf_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['f']))
            ws[ws._current_row][self.cb_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['b']))
            ws[ws._current_row][self.ca_index].fill = PatternFill(
                fill_type='gray125', end_color=self.gen_per_color(ch_dic['a']))
    def proc_fc_sheet(self, ws):
        cg_score_dic = self.parse_cg_report()
        if self.gen_flg:
            ws.append(self.fc_col_lst)
        for cell, width in zip(ws[1], pcom.rd_cfg(
                self.proj_cfg, 'vplan_column_width', ws.title)):
            if self.gen_flg:
                cell.style = 'Accent1'
            cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions[cell.column].width = width
        for index, cg_row in enumerate(ws.rows):
            if index == 0:
                continue
            cg_name = cg_row[self.fc_index].value
            if cg_name in cg_score_dic:
                per = cg_score_dic[cg_name]['per']
                cg_row[self.fs_index].value = '{0} %'.format(per)
                cg_row[self.fs_index].fill = PatternFill(
                    fill_type='gray125', end_color=self.gen_per_color(per))
            elif '::' in cg_name:
                base_cg_name = cg_name.split('::')[0]
                per = cg_score_dic[base_cg_name]['cp_dic'][cg_name]
                if cg_name in cg_score_dic[base_cg_name]['cp_dic']:
                    cg_row[self.fs_index].value = '{0} %'.format(per)
                    cg_row[self.fs_index].fill = PatternFill(
                        fill_type='gray125', end_color=self.gen_per_color(per))
                    cg_score_dic[base_cg_name]['cp_dic'].pop(cg_name)
            else:
                cg_row[self.fp_index].value = 'Out of Date'
                cg_row[self.fp_index].fill = PatternFill(
                    fill_type='gray125', end_color='FFFF0000')
            cg_pop_lst = []
            for cg_name, cg_dic in cg_score_dic.items():
                if not cg_dic['cp_dic']:
                    cg_pop_lst.append(cg_name)
            for cg_pop in cg_pop_lst:
                cg_score_dic.pop(cg_pop)
        for cg_name, cg_dic in cg_score_dic.items():
            new_line = ['']*len(self.fc_col_lst)
            new_line[self.fc_index] = cg_name
            new_line[self.fs_index] = '{0} %'.format(cg_dic['per'])
            ws.append(new_line)
            ws[ws._current_row][self.fc_index].fill = PatternFill(
                fill_type='gray125', end_color='FFFFFF00')
            ws[ws._current_row][self.fs_index].fill = PatternFill(
                fill_type='gray125',
                end_color=self.gen_per_color(cg_dic['per']))
            for cp_name, cp_per in cg_dic['cp_dic'].items():
                new_line = ['']*len(self.fc_col_lst)
                new_line[self.fc_index] = cp_name
                new_line[self.fs_index] = '{0} %'.format(cp_per)
                ws.append(new_line)
                ws[ws._current_row][self.fs_index].fill = PatternFill(
                    fill_type='gray125', end_color=self.gen_per_color(cp_per))
    def proc_vplan(self):
        if self.gen_flg:
            wb = openpyxl.Workbook()
            home_sheet = wb.active
            tc_sheet = wb.create_sheet('test_case')
            cc_sheet = wb.create_sheet('code_coverage')
            fc_sheet = wb.create_sheet('function_coverage')
        else:
            wb = openpyxl.load_workbook(self.vplan_file)
            home_sheet = wb['home']
            tc_sheet = wb['test_case']
            cc_sheet = wb['code_coverage']
            fc_sheet = wb['function_coverage']
        self.proc_tc_sheet(tc_sheet)
        self.proc_cc_sheet(cc_sheet)
        self.proc_fc_sheet(fc_sheet)
        self.proc_home_sheet(home_sheet)
        wb.save(self.vplan_file)
