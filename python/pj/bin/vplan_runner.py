"""
Author: Guanyu Yi @ CPU Verification Platform Group
Email: yigy@cpu.com.cn
Description: pj vplan sub cmd entrence and vplan flow class
"""

import os
import re
import datetime as dt
import collections
import requests
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import pcom
import env_booter

LOG = pcom.gen_logger(__name__)

class VPlanProc(object):
    """vplan flow processor and excel generator for pj"""
    def __init__(self, ced, cfg_dic, days):
        self.ced = ced
        self.cfg_dic = cfg_dic
        self.days = days
        self.gen_flg = False
        self.v_dic = {"tpn": 0, "ttn": 0, "ccs": "0", "fcp": "0"}
    @classmethod
    def gen_per_color(cls, per):
        """to generate percentage background color"""
        if not per.replace(".", "").isdigit():
            return "FFFFFFFF"
        per = float(per)
        if 0 <= per < 5:
            c_s = "FFFF0000"
        elif 5 <= per < 20:
            c_s = "FFFF3C3C"
        elif 20 <= per < 40:
            c_s = "FFFF2A00"
        elif 40 <= per < 60:
            c_s = "FFFFA500"
        elif 60 <= per < 80:
            c_s = "FFAEFF00"
        elif 80 <= per < 95:
            c_s = "FF04FF00"
        elif 95 <= per <= 100:
            c_s = "FF00FF00"
        else:
            c_s = "FFFF0000"
        return c_s
    def fill_cc_scores(self, cell, score):
        """to fill code coverage cell by using score"""
        cell.value = f"{score} %"
        cell.fill = PatternFill(fill_type="gray125", end_color=self.gen_per_color(score))
    def proc_vplan_row1(self, w_s):
        """to process vplan row1 style per sheet except home"""
        for cell, width in zip(w_s[1], pcom.rd_cfg(
                self.cfg_dic["proj"], "vplan_column_width", w_s.title)):
            if self.gen_flg:
                cell.style = "Accent1"
            cell.alignment = Alignment(wrap_text=True)
            w_s.column_dimensions[cell.column].width = width
    def parse_ch_report(self):
        """to parse coverage hierarchy report"""
        cd_rpt_file = f"{self.ced['COV_MERGE']}{os.sep}urgReport{os.sep}dashboard.txt"
        ch_rpt_file = f"{self.ced['COV_MERGE']}{os.sep}urgReport{os.sep}hierarchy.txt"
        if not os.path.isfile(cd_rpt_file):
            raise Exception(f"merged code coverage dashboard file {cd_rpt_file} is NA")
        if not os.path.isfile(ch_rpt_file):
            raise Exception(f"merged code coverage report file {ch_rpt_file} is NA")
        with open(cd_rpt_file) as rptf:
            cd_rpt_con = rptf.read()
        self.v_dic["ccs"] = re.search(
            r"Total Coverage Summary.*?(\d+\.\d+)", cd_rpt_con, flags=re.DOTALL).group(1)
        ch_score_dic = collections.OrderedDict()
        with open(ch_rpt_file) as rptf:
            ch_rpt_con = rptf.read()
        con_lst = re.findall(
            rf"{os.linesep}\s{{0,2}}-{{10,}}{os.linesep}(.*?)(?={os.linesep}\s+-{{10}}|$)",
            ch_rpt_con, flags=re.DOTALL)
        for index, con in enumerate(con_lst):
            p_str = "(top)" if index == 0 else ""
            for line in con.split(os.linesep):
                line = line.strip()
                mop = pcom.REOpter(line)
                if mop.match(re.compile(
                        r"([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+"
                        r"([\d\.\-]+)\s+([\d\.\-]+)\s+([\d\.\-]+)\s+(\w+)")):
                    ch_score_dic[f"{mop.group(8)}{p_str}"] = {
                        "s": mop.group(1), "l": mop.group(2), "c": mop.group(3), "t": mop.group(4),
                        "f": mop.group(5), "b": mop.group(6), "a": mop.group(7)}
        return ch_score_dic
    def parse_cg_report(self):
        """to parse coverage group report"""
        cg_rpt_file = f"{self.ced['COV_MERGE']}{os.sep}urgReport{os.sep}groups.txt"
        cp_rpt_file = f"{self.ced['COV_MERGE']}{os.sep}urgReport{os.sep}grpinfo.txt"
        if not os.path.isfile(cg_rpt_file):
            raise Exception(f"merged coverage groups report file {cg_rpt_file} is NA")
        cg_score_dic = collections.OrderedDict()
        with open(cg_rpt_file) as rptf:
            for line in rptf:
                line = line.strip()
                mop = pcom.REOpter(line)
                if mop.match(re.compile(r"(\d+\.\d+)\s+\d+$")):
                    self.v_dic["fcp"] = mop.group(1)
                elif mop.match(re.compile(r"(\d+\.\d+)\s+.*\w+::\w+::(\w+)")):
                    cg_score_dic[mop.group(2)] = {
                        "per": mop.group(1), "cp_dic": collections.OrderedDict()}
        if not os.path.isfile(cp_rpt_file):
            LOG.warning("merged coverage points report file %s is NA", cp_rpt_file)
            cp_rpt_con = ""
        else:
            with open(cp_rpt_file) as rptf:
                cp_rpt_con = rptf.read()
        for cg_n, cg_dic in cg_score_dic.items():
            cg_sum_con = re.search(
                rf"Summary for Group\s+(?:\w+::)+{cg_n}(.*?{os.linesep}-{{60}})",
                cp_rpt_con, flags=re.DOTALL).group(1)
            var_con, cro_con = re.search(
                rf"Variables for Group\s+(?:\w+::)+{cg_n}(.*?){os.linesep}"
                rf"Crosses for Group\s+(?:\w+::)+{cg_n}(.*?){os.linesep}-{{60}}",
                cg_sum_con, flags=re.DOTALL).groups()
            for line in var_con.split(os.linesep):
                line = line.strip()
                mop = pcom.REOpter(line)
                if mop.match(re.compile(r"(\w+)\s+(?:\d+\s+)+(\d+\.\d+)\s+(?:\d+\s+)+")):
                    cg_dic["cp_dic"][f"{cg_n}::{mop.group(1)}"] = mop.group(2)
            for line in cro_con.split(os.linesep):
                line = line.strip()
                mop = pcom.REOpter(line)
                if mop.match(re.compile(r"(\w+)\s+(?:\d+\s+)+(\d+\.\d+)\s+(?:\d+\s+)+")):
                    cg_dic["cp_dic"][f"{cg_n}::{mop.group(1)}(cross)"] = mop.group(2)
        return cg_score_dic
    @classmethod
    def clean_cg_score_dic(cls, cg_score_dic):
        """to clean no cover point items in cover group dic"""
        cg_del_lst = []
        for cg_name, cg_dic in cg_score_dic.items():
            if not cg_dic["cp_dic"]:
                cg_del_lst.append(cg_name)
        for cg_del in cg_del_lst:
            del cg_score_dic[cg_del]
    def proc_home_sheet(self, w_s):
        """to process generated vplan home sheet"""
        w_s.title = "home"
        home_row_lst = pcom.rd_cfg(self.cfg_dic["proj"], "vplan_sheets", "home")
        home_row_lst.insert(5, "")
        for index, row in enumerate(home_row_lst):
            if index == 5:
                continue
            cell = w_s[f"a{index+1}"]
            cell.value = row
            if self.gen_flg:
                cell.style = "Accent1"
            cell.alignment = Alignment(wrap_text=True)
            next_cell = w_s[f"b{index+1}"]
            if row == "Project":
                next_cell.value = self.ced["PROJ_NAME"]
            elif row == "Module Name":
                next_cell.value = self.ced["MODULE"]
            elif row == "Case Passing Rate":
                d_v = self.v_dic["tpn"]/self.v_dic["ttn"] if self.v_dic["ttn"] else 0
                cpr = str(round(100*d_v, 2))
                next_cell.value = f"{cpr} % ({self.v_dic['tpn']}/{self.v_dic['ttn']})"
                next_cell.fill = PatternFill(
                    fill_type="gray125", end_color=self.gen_per_color(cpr))
            elif row == "Code Coverage Score":
                next_cell.value = f"{self.v_dic['ccs']} %"
                next_cell.fill = PatternFill(
                    fill_type="gray125", end_color=self.gen_per_color(self.v_dic["ccs"]))
            elif row == "Function Coverage Per":
                next_cell.value = f"{self.v_dic['fcp']} %"
                next_cell.fill = PatternFill(
                    fill_type="gray125", end_color=self.gen_per_color(self.v_dic["fcp"]))
        w_s.column_dimensions["a"].width = pcom.rd_cfg(
            self.cfg_dic["proj"], "vplan_column_width", w_s.title, True)
        w_s.column_dimensions["b"].width = w_s.column_dimensions["a"].width
    def proc_tc_sheet(self, w_s):
        """to process generated vplan test case sheet"""
        query_url = "http://172.51.13.205:8000/pj_app/regr/db_query/query_case_dic/"
        query_param = {
            "date": dt.datetime.now().strftime("%Y_%m_%d"),
            "proj": self.ced["PROJ_NAME"],
            "module": self.ced["MODULE"],
            "days": self.days}
        case_pr_dic = requests.get(query_url, params=query_param).json()
        tc_col_lst = pcom.rd_cfg(self.cfg_dic["proj"], "vplan_sheets", "test_case")
        index_dic = {
            "c": tc_col_lst.index("Case Name"),
            "p": tc_col_lst.index("Priority"),
            "o": tc_col_lst.index("Owner"),
            "s": tc_col_lst.index("Status"),
            "r": tc_col_lst.index("Days"),
            "v": tc_col_lst.index("CL Ver"),
            "d": tc_col_lst.index("Description")}
        if self.gen_flg:
            w_s.append(tc_col_lst)
        self.proc_vplan_row1(w_s)
        for index, case_row in enumerate(w_s.rows):
            if index == 0:
                continue
            case_name = case_row[index_dic["c"]].value
            case_dic = case_pr_dic.get(case_name, {})
            if case_name in self.cfg_dic["case"]:
                case_row[index_dic["d"]].value = self.cfg_dic["case"][case_name][
                    "vplan_desc"].replace(os.linesep, "; ")
                case_row[index_dic["p"]].value = pcom.rd_cfg(
                    self.cfg_dic["case"], case_name, "vplan_priority", True)
                case_row[index_dic["o"]].value = pcom.rd_cfg(
                    self.cfg_dic["case"], case_name, "vplan_owner", True)
                case_row[index_dic["s"]].value = (
                    f"{case_dic.get('pr', 0.0)} % "
                    f"({case_dic.get('pn', 0)}/{case_dic.get('tn', 0)})")
                self.v_dic["tpn"] += case_dic.get("pn", 0)
                self.v_dic["ttn"] += case_dic.get("tn", 0)
                case_row[index_dic["s"]].fill = PatternFill(
                    fill_type="gray125",
                    end_color=case_dic.get("bc", "#FF0000").replace("#", "FF"))
                case_row[index_dic["r"]].value = self.days
                case_row[index_dic["v"]].value = case_dic.get("cl_range", "NA")
                del self.cfg_dic["case"][case_name]
            else:
                case_row[index_dic["p"]].value = "Out of Date"
                case_row[index_dic["p"]].fill = PatternFill(
                    fill_type="gray125", end_color="FFFF0000")
        for case_name in self.cfg_dic["case"]:
            if case_name == "DEFAULT":
                continue
            case_dic = case_pr_dic.get(case_name, {})
            new_line = [""]*len(tc_col_lst)
            new_line[index_dic["c"]] = case_name
            new_line[index_dic["d"]] = self.cfg_dic["case"][case_name][
                "vplan_desc"].replace(os.linesep, "; ")
            new_line[index_dic["p"]] = pcom.rd_cfg(
                self.cfg_dic["case"], case_name, "vplan_priority", True)
            new_line[index_dic["o"]] = pcom.rd_cfg(
                self.cfg_dic["case"], case_name, "vplan_owner", True)
            new_line[index_dic["s"]] = (
                f"{case_dic.get('pr', 0.0)} % "
                f"({case_dic.get('pn', 0)}/{case_dic.get('tn', 0)})")
            self.v_dic["tpn"] += case_dic.get("pn", 0)
            self.v_dic["ttn"] += case_dic.get("tn", 0)
            new_line[index_dic["r"]] = self.days
            new_line[index_dic["v"]] = case_dic.get("cl_range", "NA")
            w_s.append(new_line)
            w_s[w_s.max_row][index_dic["s"]].fill = PatternFill(
                fill_type="gray125",
                end_color=case_dic.get("bc", "#FF0000").replace("#", "FF"))
    def proc_cc_sheet(self, w_s):
        """to process generated vplan code coverage sheet"""
        ch_score_dic = self.parse_ch_report()
        cc_col_lst = pcom.rd_cfg(self.cfg_dic["proj"], "vplan_sheets", "code_coverage")
        index_dic = {
            "h": cc_col_lst.index("Hierarchy"),
            "p": cc_col_lst.index("Priority"),
            "s": cc_col_lst.index("Score"),
            "l": cc_col_lst.index("Line"),
            "c": cc_col_lst.index("Cond"),
            "t": cc_col_lst.index("Toggle"),
            "f": cc_col_lst.index("FSM"),
            "b": cc_col_lst.index("Branch"),
            "a": cc_col_lst.index("Assert")}
        if self.gen_flg:
            w_s.append(cc_col_lst)
        self.proc_vplan_row1(w_s)
        for index, ch_row in enumerate(w_s.rows):
            if index == 0:
                continue
            ch_name = ch_row[index_dic["h"]].value
            if ch_name in ch_score_dic:
                self.fill_cc_scores(ch_row[index_dic["s"]], ch_score_dic[ch_name]["s"])
                self.fill_cc_scores(ch_row[index_dic["l"]], ch_score_dic[ch_name]["l"])
                self.fill_cc_scores(ch_row[index_dic["c"]], ch_score_dic[ch_name]["c"])
                self.fill_cc_scores(ch_row[index_dic["t"]], ch_score_dic[ch_name]["t"])
                self.fill_cc_scores(ch_row[index_dic["f"]], ch_score_dic[ch_name]["f"])
                self.fill_cc_scores(ch_row[index_dic["b"]], ch_score_dic[ch_name]["b"])
                self.fill_cc_scores(ch_row[index_dic["a"]], ch_score_dic[ch_name]["a"])
                del ch_score_dic[ch_name]
            else:
                ch_row[index_dic["p"]].value = "Out of Date"
                ch_row[index_dic["p"]].fill = PatternFill(
                    fill_type="gray125", end_color="FFFF0000")
        for ch_name, ch_dic in ch_score_dic.items():
            new_line = [""]*len(cc_col_lst)
            new_line[index_dic["h"]] = ch_name
            new_line[index_dic["s"]] = f"{ch_dic['s']} %"
            new_line[index_dic["l"]] = f"{ch_dic['l']} %"
            new_line[index_dic["c"]] = f"{ch_dic['c']} %"
            new_line[index_dic["t"]] = f"{ch_dic['t']} %"
            new_line[index_dic["f"]] = f"{ch_dic['f']} %"
            new_line[index_dic["b"]] = f"{ch_dic['b']} %"
            new_line[index_dic["a"]] = f"{ch_dic['a']} %"
            w_s.append(new_line)
            if "(top)" in ch_name:
                w_s[w_s.max_row][index_dic["h"]].fill = PatternFill(
                    fill_type="gray125", end_color="FFFFFF00")
            w_s[w_s.max_row][index_dic["s"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["s"]))
            w_s[w_s.max_row][index_dic["l"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["l"]))
            w_s[w_s.max_row][index_dic["c"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["c"]))
            w_s[w_s.max_row][index_dic["t"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["t"]))
            w_s[w_s.max_row][index_dic["f"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["f"]))
            w_s[w_s.max_row][index_dic["b"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["b"]))
            w_s[w_s.max_row][index_dic["a"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(ch_dic["a"]))
    def proc_fc_sheet(self, w_s):
        """to process generated vplan function coverage sheet"""
        cg_score_dic = self.parse_cg_report()
        fc_col_lst = pcom.rd_cfg(self.cfg_dic["proj"], "vplan_sheets", "function_coverage")
        index_dic = {
            "c": fc_col_lst.index("Coverage Group"),
            "p": fc_col_lst.index("Priority"),
            "s": fc_col_lst.index("SNPS Cov Per")}
        if self.gen_flg:
            w_s.append(fc_col_lst)
        self.proc_vplan_row1(w_s)
        for index, cg_row in enumerate(w_s.rows):
            if index == 0:
                continue
            cg_name = cg_row[index_dic["c"]].value
            if cg_name in cg_score_dic:
                per = cg_score_dic[cg_name]["per"]
                cg_row[index_dic["s"]].value = f"{per} %"
                cg_row[index_dic["s"]].fill = PatternFill(
                    fill_type="gray125", end_color=self.gen_per_color(per))
            elif "::" in cg_name:
                base_cg_name = cg_name.split("::")[0]
                per = cg_score_dic[base_cg_name]["cp_dic"][cg_name]
                if cg_name in cg_score_dic[base_cg_name]["cp_dic"]:
                    cg_row[index_dic["s"]].value = f"{per} %"
                    cg_row[index_dic["s"]].fill = PatternFill(
                        fill_type="gray125", end_color=self.gen_per_color(per))
                    del cg_score_dic[base_cg_name]["cp_dic"][cg_name]
            else:
                cg_row[index_dic["p"]].value = "Out of Date"
                cg_row[index_dic["p"]].fill = PatternFill(
                    fill_type="gray125", end_color="FFFF0000")
            self.clean_cg_score_dic(cg_score_dic)
        for cg_name, cg_dic in cg_score_dic.items():
            new_line = [""]*len(fc_col_lst)
            new_line[index_dic["c"]] = cg_name
            new_line[index_dic["s"]] = f"{cg_dic['per']} %"
            w_s.append(new_line)
            w_s[w_s.max_row][index_dic["c"]].fill = PatternFill(
                fill_type="gray125", end_color="FFFFFF00")
            w_s[w_s.max_row][index_dic["s"]].fill = PatternFill(
                fill_type="gray125", end_color=self.gen_per_color(cg_dic["per"]))
            for cp_name, cp_per in cg_dic["cp_dic"].items():
                new_line = [""]*len(fc_col_lst)
                new_line[index_dic["c"]] = cp_name
                new_line[index_dic["s"]] = f"{cp_per} %"
                w_s.append(new_line)
                w_s[w_s.max_row][index_dic["s"]].fill = PatternFill(
                    fill_type="gray125", end_color=self.gen_per_color(cp_per))
    def proc_vplan(self):
        """top execution function"""
        vplan_file = f"{self.ced['MODULE_VPLAN']}{os.sep}{self.ced['MODULE']}_vplan.xlsx"
        self.gen_flg = False if os.path.isfile(vplan_file) else True
        if self.gen_flg:
            w_b = openpyxl.Workbook()
            home_sheet = w_b.active
            tc_sheet = w_b.create_sheet("test_case")
            cc_sheet = w_b.create_sheet("code_coverage")
            fc_sheet = w_b.create_sheet("function_coverage")
        else:
            w_b = openpyxl.load_workbook(vplan_file)
            home_sheet = w_b["home"]
            tc_sheet = w_b["test_case"]
            cc_sheet = w_b["code_coverage"]
            fc_sheet = w_b["function_coverage"]
        self.proc_tc_sheet(tc_sheet)
        self.proc_cc_sheet(cc_sheet)
        self.proc_fc_sheet(fc_sheet)
        self.proc_home_sheet(home_sheet)
        w_b.save(vplan_file)

def run_vplan(args):
    """to run vplan sub cmd"""
    if args.vplan_module and args.vplan_proc:
        ced, cfg_dic = env_booter.EnvBooter().module_env(args.vplan_module)
        VPlanProc(ced, cfg_dic, args.vplan_days).proc_vplan()
        LOG.info("processing vplan of %s module done", args.vplan_module)
    else:
        raise Exception("missing main arguments")
